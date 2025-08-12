#!/usr/bin/env python3
"""
Create test files in both .xlsx and .xls formats.
"""

import os
from openpyxl import Workbook
import xlwt

# Test data
data = [
    ['Name', 'Age', 'City', 'Score'],
    ['Alice', 25, 'New York', 95],
    ['Bob', 30, 'Los Angeles', 87],
    ['Charlie', 35, 'Chicago', 92],
    ['Diana', 28, 'Houston', 88]
]

folders = ['prev', 'new', 'template']

print("Creating test files...")

for folder in folders:
    if not os.path.exists(folder):
        os.makedirs(folder)
    
    # Create .xlsx file
    xlsx_path = os.path.join(folder, "test_data.xlsx")
    wb = Workbook()
    ws = wb.active
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(xlsx_path)
    print(f"✅ Created: {xlsx_path}")
    
    # Create .xls file
    xls_path = os.path.join(folder, "legacy_data.xls")
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet('Sheet1')
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            worksheet.write(row_idx, col_idx, value)
    workbook.save(xls_path)
    print(f"✅ Created: {xls_path}")

print("\n🎉 All test files created successfully!")
print("📁 Files in each folder:")
print("   📄 test_data.xlsx (modern Excel format)")
print("   📋 legacy_data.xls (legacy Excel format)")

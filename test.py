import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import shutil
from pathlib import Path
import numpy as np

class ExcelComparator:
    def __init__(self, base_path):
        """
        Initialize the Excel Comparator
        
        Args:
            base_path (str): Base directory containing new, prev, and template folders
        """
        self.base_path = Path(base_path)
        self.new_folder = self.base_path / "new"
        self.prev_folder = self.base_path / "prev"
        self.template_folder = self.base_path / "template"
        self.output_folder = self.base_path / "comparison_results"
        
        # Create output folder if it doesn't exist
        self.output_folder.mkdir(exist_ok=True)
        
        # Define red fill for highlighting differences
        self.red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        self.red_font = Font(color="FFFFFF", bold=True)
    
    def get_matching_files(self):
        """
        Get list of Excel files that exist in both new and prev folders
        
        Returns:
            list: List of file names present in both folders
        """
        new_files = set(f.name for f in self.new_folder.glob("*.xlsx"))
        prev_files = set(f.name for f in self.prev_folder.glob("*.xlsx"))
        template_files = set(f.name for f in self.template_folder.glob("*.xlsx"))
        
        # Get files that exist in all three folders
        common_files = new_files.intersection(prev_files).intersection(template_files)
        
        if not common_files:
            print("No matching Excel files found in all three folders (new, prev, template)")
            return []
        
        print(f"Found {len(common_files)} matching files:")
        for file in common_files:
            print(f"  - {file}")
        
        return list(common_files)
    
    def is_numeric(self, value):
        """
        Check if a value is numeric (int, float, or can be converted to float)
        
        Args:
            value: The value to check
            
        Returns:
            bool: True if numeric, False otherwise
        """
        if pd.isna(value) or value is None or value == "":
            return False
        
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False
    
    def compare_sheets(self, new_file, prev_file, template_file, output_file):
        """
        Compare all sheets in the Excel files and highlight differences
        
        Args:
            new_file (Path): Path to new Excel file
            prev_file (Path): Path to previous Excel file
            template_file (Path): Path to template Excel file
            output_file (Path): Path to output Excel file
        """
        # Copy template file to output location to preserve formatting
        shutil.copy2(template_file, output_file)
        
        # Load workbooks
        try:
            new_wb = openpyxl.load_workbook(new_file, data_only=True)
            prev_wb = openpyxl.load_workbook(prev_file, data_only=True)
            output_wb = openpyxl.load_workbook(output_file)
        except Exception as e:
            print(f"Error loading workbooks: {e}")
            return
        
        # Get common sheet names
        new_sheets = set(new_wb.sheetnames)
        prev_sheets = set(prev_wb.sheetnames)
        template_sheets = set(output_wb.sheetnames)
        
        common_sheets = new_sheets.intersection(prev_sheets).intersection(template_sheets)
        
        if not common_sheets:
            print(f"No common sheets found in {new_file.name}")
            return
        
        print(f"\nProcessing {len(common_sheets)} sheets in {new_file.name}:")
        
        differences_found = False
        
        for sheet_name in common_sheets:
            print(f"  Comparing sheet: {sheet_name}")
            
            try:
                new_sheet = new_wb[sheet_name]
                prev_sheet = prev_wb[sheet_name]
                output_sheet = output_wb[sheet_name]
                
                sheet_differences = self.compare_single_sheet(
                    new_sheet, prev_sheet, output_sheet, sheet_name
                )
                
                if sheet_differences:
                    differences_found = True
                    
            except Exception as e:
                print(f"    Error comparing sheet {sheet_name}: {e}")
                continue
        
        # Save the output file
        try:
            output_wb.save(output_file)
            if differences_found:
                print(f"  ✓ Differences found and highlighted in: {output_file.name}")
            else:
                print(f"  ✓ No differences found in: {output_file.name}")
        except Exception as e:
            print(f"  ✗ Error saving output file: {e}")
        finally:
            new_wb.close()
            prev_wb.close()
            output_wb.close()
    
    def compare_single_sheet(self, new_sheet, prev_sheet, output_sheet, sheet_name):
        """
        Compare a single sheet and highlight differences
        
        Args:
            new_sheet: New sheet object
            prev_sheet: Previous sheet object
            output_sheet: Output sheet object
            sheet_name: Name of the sheet
            
        Returns:
            bool: True if differences were found, False otherwise
        """
        differences_found = False
        differences_count = 0
        
        # Get the maximum row and column from both sheets
        max_row = max(new_sheet.max_row, prev_sheet.max_row)
        max_col = max(new_sheet.max_column, prev_sheet.max_column)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    # Get cell values
                    new_cell = new_sheet.cell(row=row, column=col)
                    prev_cell = prev_sheet.cell(row=row, column=col)
                    output_cell = output_sheet.cell(row=row, column=col)
                    
                    new_value = new_cell.value
                    prev_value = prev_cell.value
                    
                    # Skip if both cells are empty
                    if (new_value is None or new_value == "") and (prev_value is None or prev_value == ""):
                        continue
                    
                    # Check if values are different
                    if new_value != prev_value:
                        # Check if both values are numeric
                        if self.is_numeric(new_value) and self.is_numeric(prev_value):
                            try:
                                new_num = float(new_value) if new_value is not None else 0
                                prev_num = float(prev_value) if prev_value is not None else 0
                                difference = new_num - prev_num
                                
                                # Update the output cell with the difference
                                output_cell.value = difference
                                
                                # Apply red highlighting
                                output_cell.fill = self.red_fill
                                output_cell.font = self.red_font
                                
                                differences_found = True
                                differences_count += 1
                                
                            except (ValueError, TypeError):
                                # If conversion fails, treat as text difference
                                output_cell.value = new_value
                                output_cell.fill = self.red_fill
                                output_cell.font = self.red_font
                                differences_found = True
                                differences_count += 1
                        
                        elif self.is_numeric(new_value) and not self.is_numeric(prev_value):
                            # New value is numeric, old is not
                            try:
                                new_num = float(new_value)
                                output_cell.value = new_num  # Show the new numeric value
                                output_cell.fill = self.red_fill
                                output_cell.font = self.red_font
                                differences_found = True
                                differences_count += 1
                            except (ValueError, TypeError):
                                pass
                        
                        elif not self.is_numeric(new_value) and self.is_numeric(prev_value):
                            # Old value was numeric, new is not
                            output_cell.value = new_value  # Show the new non-numeric value
                            output_cell.fill = self.red_fill
                            output_cell.font = self.red_font
                            differences_found = True
                            differences_count += 1
                        
                        else:
                            # Both are non-numeric but different
                            output_cell.value = new_value
                            output_cell.fill = self.red_fill
                            output_cell.font = self.red_font
                            differences_found = True
                            differences_count += 1
                
                except Exception as e:
                    print(f"    Error processing cell {row},{col}: {e}")
                    continue
        
        if differences_count > 0:
            print(f"    Found {differences_count} differences in sheet '{sheet_name}'")
        
        return differences_found
    
    def run_comparison(self):
        """
        Run the complete comparison process for all matching files
        """
        print("Starting Excel Comparison Process...")
        print("="*50)
        
        matching_files = self.get_matching_files()
        
        if not matching_files:
            print("No files to compare. Exiting.")
            return
        
        print(f"\nProcessing {len(matching_files)} files...")
        print("="*50)
        
        for file_name in matching_files:
            print(f"\nProcessing file: {file_name}")
            print("-" * 40)
            
            new_file = self.new_folder / file_name
            prev_file = self.prev_folder / file_name
            template_file = self.template_folder / file_name
            
            # Create output filename with timestamp
            name_parts = file_name.rsplit('.', 1)
            output_name = f"{name_parts[0]}_COMPARISON.xlsx"
            output_file = self.output_folder / output_name
            
            try:
                self.compare_sheets(new_file, prev_file, template_file, output_file)
            except Exception as e:
                print(f"  ✗ Error processing {file_name}: {e}")
                continue
        
        print("\n" + "="*50)
        print("Comparison process completed!")
        print(f"Results saved in: {self.output_folder}")
        print("\nOutput files contain:")
        print("- Original formatting from template files")
        print("- Differences highlighted in RED")
        print("- For numeric differences: shows (new value - old value)")
        print("- For non-numeric differences: shows the new value")


def main():
    """
    Main function to run the Excel comparison
    """
    # Get the current directory (where the script is located)
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    print("Excel File Comparison Tool")
    print("="*30)
    print(f"Working directory: {current_dir}")
    print()
    
    # Initialize the comparator
    comparator = ExcelComparator(current_dir)
    
    # Run the comparison
    comparator.run_comparison()


if __name__ == "__main__":
    main()
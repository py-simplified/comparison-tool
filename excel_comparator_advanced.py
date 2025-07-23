import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import shutil
from pathlib import Path
import numpy as np
from datetime import datetime
import json
import getpass
import hashlib

class ExcelComparatorAdvanced:
    def __init__(self, base_path):
        """
        Initialize the Advanced Excel Comparator
        
        Args:
            base_path (str): Base directory containing new, prev, and template folders
        """
        self.base_path = Path(base_path)
        self.new_folder = self.base_path / "new"
        self.prev_folder = self.base_path / "prev"
        self.template_folder = self.base_path / "template"
        self.output_folder = self.base_path / "comparison_results"
        
        # Create output folder if it doesn't exist
        self.output_folder.mkdir(exist_ok=True)
        
        # Define red fill for highlighting differences
        self.red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        self.red_font = Font(color="FFFFFF", bold=True)
        
        # Initialize summary data
        self.summary_data = {
            "timestamp": datetime.now().isoformat(),
            "files_processed": [],
            "total_differences": 0,
            "errors": []
        }
        
        # Default password (hashed) - Change this in production
        # Default password is "1234" (hashed with SHA-256)
        self.password_hash = "9af15b336e6a9619928537df30b2e6a2376569fcf9d7e773eccede65606529a0"
    
    def verify_password(self):
        """
        Verify user password before allowing access to the tool
        
        Returns:
            bool: True if password is correct, False otherwise
        """
        max_attempts = 3
        attempts = 0
        
        print("🔒 Advanced Excel Comparison Tool - Password Protection")
        print("=" * 55)
        
        while attempts < max_attempts:
            try:
                # Use getpass to hide password input
                password = getpass.getpass(f"Enter 4-digit password (Attempt {attempts + 1}/{max_attempts}): ")
                
                # Validate password format
                if not password.isdigit() or len(password) != 4:
                    print("❌ Invalid format! Password must be exactly 4 digits.")
                    attempts += 1
                    continue
                
                # Hash the entered password
                password_hash = hashlib.sha256(password.encode()).hexdigest()
                
                # Compare with stored hash
                if password_hash == self.password_hash:
                    print("✅ Password correct! Access granted.")
                    print()
                    return True
                else:
                    print("❌ Incorrect password!")
                    attempts += 1
                    
            except KeyboardInterrupt:
                print("\n🚫 Operation cancelled by user.")
                return False
            except Exception as e:
                print(f"❌ Error during password verification: {e}")
                attempts += 1
        
        print(f"🚫 Maximum attempts ({max_attempts}) exceeded. Access denied.")
        return False
    
    def change_password(self):
        """
        Allow changing the password (admin function)
        This should be called separately for security setup
        """
        print("🔧 Password Change Utility")
        print("=" * 30)
        
        # Verify current password first
        if not self.verify_password():
            return False
        
        while True:
            new_password = getpass.getpass("Enter new 4-digit password: ")
            
            if not new_password.isdigit() or len(new_password) != 4:
                print("❌ Invalid format! Password must be exactly 4 digits.")
                continue
            
            confirm_password = getpass.getpass("Confirm new password: ")
            
            if new_password != confirm_password:
                print("❌ Passwords don't match! Please try again.")
                continue
            
            # Hash and store new password
            new_hash = hashlib.sha256(new_password.encode()).hexdigest()
            print(f"✅ New password hash: {new_hash}")
            print("📝 Update the password_hash variable in the code with this hash.")
            return True
    
    def get_matching_files(self):
        """
        Get list of Excel files that exist in both new and prev folders
        
        Returns:
            list: List of file names present in both folders
        """
        new_files = set(f.name for f in self.new_folder.glob("*.xlsx"))
        prev_files = set(f.name for f in self.prev_folder.glob("*.xlsx"))
        template_files = set(f.name for f in self.template_folder.glob("*.xlsx"))
        
        # Get files that exist in all three folders
        common_files = new_files.intersection(prev_files).intersection(template_files)
        
        if not common_files:
            print("No matching Excel files found in all three folders (new, prev, template)")
            return []
        
        print(f"Found {len(common_files)} matching files:")
        for file in common_files:
            print(f"  - {file}")
        
        return list(common_files)
    
    def is_numeric(self, value):
        """
        Check if a value is numeric (int, float, or can be converted to float)
        
        Args:
            value: The value to check
            
        Returns:
            bool: True if numeric, False otherwise
        """
        if pd.isna(value) or value is None or value == "":
            return False
        
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False
    
    def compare_sheets(self, new_file, prev_file, template_file, output_file):
        """
        Compare all sheets in the Excel files and highlight differences
        
        Args:
            new_file (Path): Path to new Excel file
            prev_file (Path): Path to previous Excel file
            template_file (Path): Path to template Excel file
            output_file (Path): Path to output Excel file
            
        Returns:
            dict: Summary of comparison results
        """
        # Copy template file to output location to preserve formatting
        shutil.copy2(template_file, output_file)
        
        file_summary = {
            "filename": new_file.name,
            "sheets_processed": 0,
            "total_differences": 0,
            "sheet_details": {},
            "errors": []
        }
        
        # Load workbooks
        try:
            new_wb = openpyxl.load_workbook(new_file, data_only=True)
            prev_wb = openpyxl.load_workbook(prev_file, data_only=True)
            output_wb = openpyxl.load_workbook(output_file)
        except Exception as e:
            error_msg = f"Error loading workbooks: {e}"
            print(error_msg)
            file_summary["errors"].append(error_msg)
            return file_summary
        
        # Get common sheet names
        new_sheets = set(new_wb.sheetnames)
        prev_sheets = set(prev_wb.sheetnames)
        template_sheets = set(output_wb.sheetnames)
        
        common_sheets = new_sheets.intersection(prev_sheets).intersection(template_sheets)
        
        if not common_sheets:
            error_msg = f"No common sheets found in {new_file.name}"
            print(error_msg)
            file_summary["errors"].append(error_msg)
            return file_summary
        
        print(f"\nProcessing {len(common_sheets)} sheets in {new_file.name}:")
        
        differences_found = False
        
        for sheet_name in common_sheets:
            print(f"  Comparing sheet: {sheet_name}")
            
            try:
                new_sheet = new_wb[sheet_name]
                prev_sheet = prev_wb[sheet_name]
                output_sheet = output_wb[sheet_name]
                
                sheet_result = self.compare_single_sheet(
                    new_sheet, prev_sheet, output_sheet, sheet_name
                )
                
                file_summary["sheet_details"][sheet_name] = sheet_result
                file_summary["total_differences"] += sheet_result["differences_count"]
                file_summary["sheets_processed"] += 1
                
                if sheet_result["differences_found"]:
                    differences_found = True
                    
            except Exception as e:
                error_msg = f"Error comparing sheet {sheet_name}: {e}"
                print(f"    {error_msg}")
                file_summary["errors"].append(error_msg)
                continue
        
        # Save the output file
        try:
            output_wb.save(output_file)
            if differences_found:
                print(f"  ✓ {file_summary['total_differences']} differences found and highlighted in: {output_file.name}")
            else:
                print(f"  ✓ No differences found in: {output_file.name}")
                
            file_summary["status"] = "completed"
            
        except Exception as e:
            error_msg = f"Error saving output file: {e}"
            print(f"  ✗ {error_msg}")
            file_summary["errors"].append(error_msg)
            file_summary["status"] = "failed"
        finally:
            new_wb.close()
            prev_wb.close()
            output_wb.close()
            
        return file_summary
    
    def compare_single_sheet(self, new_sheet, prev_sheet, output_sheet, sheet_name):
        """
        Compare a single sheet and highlight differences
        
        Args:
            new_sheet: New sheet object
            prev_sheet: Previous sheet object
            output_sheet: Output sheet object
            sheet_name: Name of the sheet
            
        Returns:
            dict: Summary of sheet comparison
        """
        sheet_result = {
            "differences_found": False,
            "differences_count": 0,
            "numeric_differences": 0,
            "text_differences": 0,
            "cells_compared": 0
        }
        
        # Get the maximum row and column from both sheets
        max_row = max(new_sheet.max_row, prev_sheet.max_row)
        max_col = max(new_sheet.max_column, prev_sheet.max_column)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    # Get cell values
                    new_cell = new_sheet.cell(row=row, column=col)
                    prev_cell = prev_sheet.cell(row=row, column=col)
                    output_cell = output_sheet.cell(row=row, column=col)
                    
                    new_value = new_cell.value
                    prev_value = prev_cell.value
                    
                    sheet_result["cells_compared"] += 1
                    
                    # Skip if both cells are empty
                    if (new_value is None or new_value == "") and (prev_value is None or prev_value == ""):
                        continue
                    
                    # Check if values are different
                    if new_value != prev_value:
                        # Check if both values are numeric
                        if self.is_numeric(new_value) and self.is_numeric(prev_value):
                            try:
                                new_num = float(new_value) if new_value is not None else 0
                                prev_num = float(prev_value) if prev_value is not None else 0
                                difference = new_num - prev_num
                                
                                # Update the output cell with the difference
                                output_cell.value = difference
                                
                                # Apply red highlighting
                                output_cell.fill = self.red_fill
                                output_cell.font = self.red_font
                                
                                sheet_result["differences_found"] = True
                                sheet_result["differences_count"] += 1
                                sheet_result["numeric_differences"] += 1
                                
                            except (ValueError, TypeError):
                                # If conversion fails, treat as text difference
                                output_cell.value = new_value
                                output_cell.fill = self.red_fill
                                output_cell.font = self.red_font
                                sheet_result["differences_found"] = True
                                sheet_result["differences_count"] += 1
                                sheet_result["text_differences"] += 1
                        
                        elif self.is_numeric(new_value) and not self.is_numeric(prev_value):
                            # New value is numeric, old is not
                            try:
                                new_num = float(new_value)
                                output_cell.value = new_num  # Show the new numeric value
                                output_cell.fill = self.red_fill
                                output_cell.font = self.red_font
                                sheet_result["differences_found"] = True
                                sheet_result["differences_count"] += 1
                                sheet_result["numeric_differences"] += 1
                            except (ValueError, TypeError):
                                pass
                        
                        elif not self.is_numeric(new_value) and self.is_numeric(prev_value):
                            # Old value was numeric, new is not
                            output_cell.value = new_value  # Show the new non-numeric value
                            output_cell.fill = self.red_fill
                            output_cell.font = self.red_font
                            sheet_result["differences_found"] = True
                            sheet_result["differences_count"] += 1
                            sheet_result["text_differences"] += 1
                        
                        else:
                            # Both are non-numeric but different
                            output_cell.value = new_value
                            output_cell.fill = self.red_fill
                            output_cell.font = self.red_font
                            sheet_result["differences_found"] = True
                            sheet_result["differences_count"] += 1
                            sheet_result["text_differences"] += 1
                
                except Exception as e:
                    print(f"    Error processing cell {row},{col}: {e}")
                    continue
        
        if sheet_result["differences_count"] > 0:
            print(f"    Found {sheet_result['differences_count']} differences in sheet '{sheet_name}' " +
                  f"({sheet_result['numeric_differences']} numeric, {sheet_result['text_differences']} text)")
        
        return sheet_result
    
    def generate_summary_report(self):
        """
        Generate a detailed summary report of the comparison process
        """
        summary_file = self.output_folder / "comparison_summary.json"
        report_file = self.output_folder / "comparison_report.txt"
        
        # Save JSON summary
        with open(summary_file, 'w') as f:
            json.dump(self.summary_data, f, indent=2)
        
        # Generate text report
        with open(report_file, 'w') as f:
            f.write("EXCEL COMPARISON SUMMARY REPORT\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Comparison completed at: {self.summary_data['timestamp']}\n")
            f.write(f"Total files processed: {len(self.summary_data['files_processed'])}\n")
            f.write(f"Total differences found: {self.summary_data['total_differences']}\n\n")
            
            if self.summary_data['errors']:
                f.write("ERRORS ENCOUNTERED:\n")
                f.write("-" * 20 + "\n")
                for error in self.summary_data['errors']:
                    f.write(f"- {error}\n")
                f.write("\n")
            
            f.write("FILE DETAILS:\n")
            f.write("-" * 20 + "\n")
            for file_data in self.summary_data['files_processed']:
                f.write(f"\nFile: {file_data['filename']}\n")
                f.write(f"  Status: {file_data.get('status', 'unknown')}\n")
                f.write(f"  Sheets processed: {file_data['sheets_processed']}\n")
                f.write(f"  Total differences: {file_data['total_differences']}\n")
                
                if file_data['sheet_details']:
                    f.write("  Sheet breakdown:\n")
                    for sheet_name, details in file_data['sheet_details'].items():
                        f.write(f"    {sheet_name}: {details['differences_count']} differences " +
                               f"({details['numeric_differences']} numeric, {details['text_differences']} text)\n")
                
                if file_data['errors']:
                    f.write("  Errors:\n")
                    for error in file_data['errors']:
                        f.write(f"    - {error}\n")
        
        print(f"\nSummary report saved to: {report_file}")
        print(f"Detailed JSON data saved to: {summary_file}")
    
    def run_comparison(self):
        """
        Run the complete comparison process for all matching files
        """
        print("Starting Excel Comparison Process...")
        print("="*50)
        
        matching_files = self.get_matching_files()
        
        if not matching_files:
            print("No files to compare. Exiting.")
            return
        
        print(f"\nProcessing {len(matching_files)} files...")
        print("="*50)
        
        for file_name in matching_files:
            print(f"\nProcessing file: {file_name}")
            print("-" * 40)
            
            new_file = self.new_folder / file_name
            prev_file = self.prev_folder / file_name
            template_file = self.template_folder / file_name
            
            # Create output filename with timestamp
            name_parts = file_name.rsplit('.', 1)
            output_name = f"{name_parts[0]}_COMPARISON.xlsx"
            output_file = self.output_folder / output_name
            
            try:
                file_summary = self.compare_sheets(new_file, prev_file, template_file, output_file)
                self.summary_data['files_processed'].append(file_summary)
                self.summary_data['total_differences'] += file_summary['total_differences']
                
                # Add any file errors to global errors
                if file_summary['errors']:
                    self.summary_data['errors'].extend(file_summary['errors'])
                    
            except Exception as e:
                error_msg = f"Error processing {file_name}: {e}"
                print(f"  ✗ {error_msg}")
                self.summary_data['errors'].append(error_msg)
                continue
        
        print("\n" + "="*50)
        print("Comparison process completed!")
        print(f"Results saved in: {self.output_folder}")
        print(f"Total differences found: {self.summary_data['total_differences']}")
        print("\nOutput files contain:")
        print("- Original formatting from template files")
        print("- Differences highlighted in RED")
        print("- For numeric differences: shows (new value - old value)")
        print("- For non-numeric differences: shows the new value")
        
        # Generate summary report
        self.generate_summary_report()


def main():
    """
    Main function to run the Excel comparison
    """
    # Get the current directory (where the script is located)
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    print("Advanced Excel File Comparison Tool")
    print("="*35)
    print(f"Working directory: {current_dir}")
    print()
    
    # Initialize the comparator
    comparator = ExcelComparatorAdvanced(current_dir)
    
    # Check for password change request
    if len(os.sys.argv) > 1 and os.sys.argv[1] == "--change-password":
        comparator.change_password()
        return
    
    # Verify password before proceeding
    if not comparator.verify_password():
        print("🚫 Access denied. Exiting...")
        input("Press Enter to exit...")
        return
    
    # Run the comparison
    comparator.run_comparison()


if __name__ == "__main__":
    main()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os

def create_test_files():
    """Create test files in both .xlsx and .xls formats"""
    
    # Sample data
    base_data = {
        'Account': ['Cash', 'Receivables', 'Inventory', 'Equipment'],
        'Q1': [50000, 120000, 80000, 200000],
        'Q2': [55000, 115000, 85000, 200000],
        'Status': ['Active', 'Active', 'Active', 'Active']
    }

    new_data = {
        'Account': ['Cash', 'Receivables', 'Inventory', 'Equipment'],
        'Q1': [52000, 120000, 82000, 200000],  # Changed values
        'Q2': [57000, 118000, 85000, 205000],  # Changed values
        'Status': ['Active', 'Active', 'Review', 'Active']  # Changed status
    }

    def create_xlsx_file(data, folder, filename):
        """Create .xlsx file using openpyxl"""
        os.makedirs(folder, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'
        
        df = pd.DataFrame(data)
        
        # Headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Data
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(f'{folder}/{filename}')
        print(f'Created .xlsx: {folder}/{filename}')

    def create_xls_file(data, folder, filename):
        """Create .xls file using pandas"""
        os.makedirs(folder, exist_ok=True)
        df = pd.DataFrame(data)
        
        # Create .xls file using xlwt engine
        filepath = f'{folder}/{filename}'
        df.to_excel(filepath, index=False, engine='xlwt')
        print(f'Created .xls: {filepath}')

    # Create .xlsx test files
    print("Creating .xlsx test files...")
    create_xlsx_file(base_data, 'prev', 'test_data.xlsx')
    create_xlsx_file(base_data, 'template', 'test_data.xlsx')
    create_xlsx_file(new_data, 'new', 'test_data.xlsx')

    # Create .xls test files
    print("Creating .xls test files...")
    try:
        create_xls_file(base_data, 'prev', 'legacy_data.xls')
        create_xls_file(base_data, 'template', 'legacy_data.xls')
        create_xls_file(new_data, 'new', 'legacy_data.xls')
    except Exception as e:
        print(f"Note: Could not create .xls files: {e}")
        print("Install xlwt with: pip install xlwt")

    print("\n✅ Test files created!")
    print("Files:")
    print("  📄 test_data.xlsx (in all folders) - .xlsx format")
    print("  📋 legacy_data.xls (in all folders) - .xls format")

if __name__ == "__main__":
    create_test_files()

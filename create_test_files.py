#!/usr/bin/env python3
"""
Create test files in both .xlsx and .xls formats for testing the Excel comparison tool.
"""

import os
from openpyxl import Workbook
import xlwt

def create_test_files():
    """Create test files in both .xlsx and .xls formats."""
    
    # Define test folders
    folders = ['prev', 'new', 'template']
    
    # Sample data for testing
    test_data = [
        ['Name', 'Age', 'City', 'Score'],
        ['Alice', 25, 'New York', 95],
        ['Bob', 30, 'Los Angeles', 87],
        ['Charlie', 35, 'Chicago', 92],
        ['Diana', 28, 'Houston', 88]
    ]
    
    # Create .xlsx test files using openpyxl
    print("Creating .xlsx test files...")
    for folder in folders:
        if not os.path.exists(folder):
            os.makedirs(folder)
        
        xlsx_path = os.path.join(folder, "test_data.xlsx")
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Write data
        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(xlsx_path)
        print(f"Created .xlsx: {xlsx_path}")
    
    # Create .xls test files using xlwt directly
    print("Creating .xls test files...")
    for folder in folders:
        xls_path = os.path.join(folder, "legacy_data.xls")
        
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet('Sheet1')
        
        # Write data
        for row_idx, row_data in enumerate(test_data):
            for col_idx, value in enumerate(row_data):
                worksheet.write(row_idx, col_idx, value)
        
        workbook.save(xls_path)
        print(f"Created .xls: {xls_path}")
    
    print("\n✅ Test files created!")
    print("Files:")
    print("  📄 test_data.xlsx (in all folders) - .xlsx format")
    print("  📋 legacy_data.xls (in all folders) - .xls format")
    return True

if __name__ == "__main__":
    create_test_files()
